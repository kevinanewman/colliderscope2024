# This Python file uses the following encoding: utf-8
import sys, os

import pandas as pd

from PySide6.QtWidgets import QApplication, QWidget, QFileDialog, QMessageBox, QTableWidgetItem

# from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QDialog, QTableWidget,
#                                QLabel, QGraphicsScene, QGraphicsWidget,
#                                QGraphicsProxyWidget, QSizePolicy, QVBoxLayout, QHBoxLayout, QCheckBox,)


# Important:
# You need to run the following command to generate the ui_form.py file
#     pyside6-uic import_tab.ui -o ui_import_tab.py, or
#     pyside2-uic import_tab.ui -o ui_import_tab.py
from import_tab.ui_import_tab import Ui_ImportTabWidget

def set_tab_by_name(tabWidget, tab_name):
    for i in range(tabWidget.count()):
        if tabWidget.tabText(i) == tab_name:
            tabWidget.setCurrentIndex(i)


def two_column_tableWidget_to_dict(table_widget):
    """
    Convert a QTableWidget (instance of a table widget in PyQt) to a dictionary using data from its first
    two columns.

    Parameters:
        table_widget (QTableWidget): The instance of the table widget.

    Returns:
        dict: A dictionary where keys and values are textual data from the first and second
              columns of each row in the table_widget. If a row doesn't have items in both columns,
              it will be ignored.

    """
    options_dict = dict()

    for r in range(table_widget.rowCount()):
        parameter = table_widget.item(r, 0)
        parameter_value = table_widget.item(r, 1)

        if (parameter is not None and
                parameter.text() != '' and
                parameter_value is not None and
                parameter_value.text() != ''):
            if parameter_value.text().lower() == 'none':
                options_dict[parameter.text()] = None
            elif parameter_value.text() == "''":
                options_dict[parameter.text()] = ''
            elif parameter_value.text().lower() == 'false':
                options_dict[parameter.text()] = False
            elif parameter_value.text().lower() == 'true':
                options_dict[parameter.text()] = True
            else:
                try:
                    options_dict[parameter.text()] = eval(parameter_value.text())
                except:
                    options_dict[parameter.text()] = parameter_value.text()

    return options_dict


def file_dialog(file_pathname, file_type_filter, file_dialog_title):
    """
    Opens a file dialog to select a file with extension options.

    :param file_pathname: Default file name
    :param file_type_filter: Specifies extension filter type
    :param file_dialog_title: Title for dialog box
    :return: User selected file name, Echo file_type, Echo file_dialog_title
    """

    dialog = QFileDialog()
    dialog.selectFile(file_pathname)
    dialog.setWindowTitle(file_dialog_title)
    dialog.setFileMode(QFileDialog.ExistingFile)
    # dialog.setFileMode(QFileDialog.AnyFile)
    if type(file_type_filter) is not list:
        dialog.setNameFilters([file_type_filter])
    else:
        dialog.setNameFilters(file_type_filter)
    dialog.setViewMode(QFileDialog.Detail)
    if dialog.exec():
        file_pathname = dialog.selectedFiles()
        file_pathname = str(file_pathname)[2:-2]
        return file_pathname
    else:
        file_pathname = ""
        return file_pathname


def get_unitized_columns(filename, sheet_name=None, ignore_units=(), encoding='utf-8', header_row=0, units_row=None,
                         skiprows=None, delimiter=None):
    """
    Combine column labels and units row into a single combined string to identify the column.

    Args:
        filename (str): name of the file to read
        sheet_name (str): for reading a particular Excel sheet, or ``None`` for CSV files
        ignore_units (list of str): unit values to ignore, default ``Text``
        encoding (str): file encoding (decoding) method name
        header_row (int): the header row number
        units_row (int): the unit row number, if any

    Returns:
        List of combined column headers and units as in ``['ColumnName_units', ...]``

    """
    num_rows = 100  # may have to read a chunk of row2 if there are partial columns...??

    if sheet_name:
        columns = pd.read_excel(filename, header=None, nrows=num_rows, skiprows=list(range(0, header_row)),
                                sheet_name=sheet_name)

        if units_row is not None:
            units = pd.read_excel(filename, header=None, nrows=num_rows, skiprows=list(range(0, units_row)),
                                  sheet_name=sheet_name)
        else:
            units = pd.DataFrame({'units': [''] * columns.shape[1]}).transpose()
    else:
        columns = pd.read_csv(filename, header=None, nrows=num_rows, skiprows=list(range(0, header_row)),
                              encoding=encoding, encoding_errors='strict', on_bad_lines='skip', delimiter=delimiter)

        if units_row is not None:
            units = pd.read_csv(filename, header=None, nrows=num_rows, skiprows=list(range(0, units_row)),
                                encoding=encoding, encoding_errors='strict', on_bad_lines='skip', delimiter=delimiter)

            # pad units with blanks if fewer units than columns...
            for i in range(max(0, len(columns.values[0]) - len(units.values[0]))):
                units[' ' * i] = ''

        else:  # blank units
            units = pd.DataFrame({'units': [''] * columns.shape[1]}).transpose()

    unitized_columns = []

    for col, unit in zip(columns.values[0], units.values[0]):
        if unit not in ignore_units and pd.notna(unit) and units_row is not None:
            if str(unit).strip() != '':
                unitized_columns.append('%s_%s' % (str(col).strip(), str(unit).strip()))
            else:
                unitized_columns.append('%s' % str(col).strip())
        else:
            unitized_columns.append('%s' % str(col).strip())

    # deal with dupicates!
    sanitized_columns = []
    raw_columns = []

    for col in unitized_columns:
        raw_columns.append(col)
        if col not in sanitized_columns:
            sanitized_columns.append(col)
        else:
            sanitized_columns.append('%s:%d' % (col, raw_columns.count(col)))

    return sanitized_columns


def import_csv_header_row_spinBox_stepBy(step):
    mainwindow.ui.import_csv_header_row_spinBox.default_stepBy(step)
    mainwindow.load_file_preview()


def import_csv_units_row_spinBox_stepBy(step):
    mainwindow.ui.import_csv_units_row_spinBox.default_stepBy(step)
    mainwindow.load_file_preview()


def import_excel_header_row_spinBox_stepBy(step):
    mainwindow.ui.import_excel_header_row_spinBox.default_stepBy(step)
    mainwindow.load_file_preview()


def import_excel_units_row_spinBox_stepBy(step):
    mainwindow.ui.import_excel_units_row_spinBox.default_stepBy(step)
    mainwindow.load_file_preview()


def preview_size_spinBox_stepBy(step):
    mainwindow.ui.preview_size_spinBox.default_stepBy(step)
    mainwindow.load_file_preview()


class ImportTabWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_ImportTabWidget()
        self.ui.setupUi(self)

        self.ui.file_import_browse_pushButton.setFocus()

        self.import_csv_options_dict = dict()
        self.import_excel_options_dict = dict()
        self.ui.import_csv_parameter_tableWidget.horizontalHeader().setMinimumHeight(25)
        self.ui.import_excel_parameter_tableWidget.horizontalHeader().setMinimumHeight(25)

        # intercept import spinBox button step events to trigger preview update without using valueChanged() signal
        self.ui.import_csv_header_row_spinBox.default_stepBy = self.ui.import_csv_header_row_spinBox.stepBy
        self.ui.import_csv_header_row_spinBox.stepBy = import_csv_header_row_spinBox_stepBy

        self.ui.import_csv_units_row_spinBox.default_stepBy = self.ui.import_csv_units_row_spinBox.stepBy
        self.ui.import_csv_units_row_spinBox.stepBy = import_csv_units_row_spinBox_stepBy

        self.ui.import_excel_header_row_spinBox.default_stepBy = self.ui.import_excel_header_row_spinBox.stepBy
        self.ui.import_excel_header_row_spinBox.stepBy = import_excel_header_row_spinBox_stepBy

        self.ui.import_excel_units_row_spinBox.default_stepBy = self.ui.import_excel_units_row_spinBox.stepBy
        self.ui.import_excel_units_row_spinBox.stepBy = import_excel_units_row_spinBox_stepBy

        self.ui.preview_size_spinBox.default_stepBy = self.ui.preview_size_spinBox.stepBy
        self.ui.preview_size_spinBox.stepBy = preview_size_spinBox_stepBy

    def generic_slot(self):
        print('generic_slot')
        pass

    def try_readline(self, f_read, stop, i):
        try:
            line = f_read.readline()
        except Exception as e:
            line = str(e)
            stop = True

        if line:
            self.ui.file_preview_tableWidget.insertRow(self.ui.file_preview_tableWidget.rowCount())
            # label rows from 0 and not 1, to be compatible with import settings like header_row:
            self.ui.file_preview_tableWidget.setVerticalHeaderItem(i, QTableWidgetItem(str(i)))
            self.ui.file_preview_tableWidget.setItem(i-1, 1, QTableWidgetItem('%s' % line.rstrip()))
        else:
            stop = True

        return stop

    def import_filebrowse(self):
        self.input_file_pathname = file_dialog('', '*.*', 'file_dialog_title')
        self.ui.filepathname_lineEdit.setText(self.input_file_pathname)
        self.ui.import_excel_sheet_comboBox.clear()
        self.handle_import_button_enables()

        self.load_file_preview('', self.input_file_pathname)

    def handle_import_button_enables(self):
        if self.ui.filepathname_lineEdit.text():
            self.ui.file_import_tabWidget.setEnabled(True)
            self.ui.row_groupBox.setEnabled(True)
            self.ui.column_groupBox.setEnabled(True)
        else:
            self.ui.file_import_tabWidget.setEnabled(False)
            self.ui.row_groupBox.setEnabled(False)
            self.ui.column_groupBox.setEnabled(False)

    def filepathname_changed(self):
        self.load_file_preview('', self.ui.filepathname_lineEdit.text())
        self.handle_import_button_enables()

    def load_file_preview(self, qstring='', file_pathname=None):
        print('load_file_preview')

        self.ui.file_import_browse_pushButton.clearFocus()

        if self.ui.preview_size_checkBox.isChecked():
            self.ui.preview_size_spinBox.setEnabled(True)
            num_preview_rows = self.ui.preview_size_spinBox.value()
        else:
            self.ui.preview_size_spinBox.setEnabled(False)
            num_preview_rows = None

        if file_pathname is None:
            file_pathname = self.ui.filepathname_lineEdit.text()

        if file_pathname:
            if 'xls' in file_pathname.split('.')[-1]:
                self.ui.import_excel_units_row_spinBox.setEnabled(self.ui.import_excel_units_row_checkBox.isChecked())

                set_tab_by_name(self.ui.file_import_tabWidget, 'Excel')

                self.ui.import_excel_pushButton.setFocus()
                self.ui.import_csv_pushButton.clearFocus()

                df = self.import_excel_file(preview=True, nrows=num_preview_rows)
                if df is not None:
                    self.preview_dataframe(df)

            else:
                self.ui.import_csv_units_row_spinBox.setEnabled(self.ui.import_csv_units_row_checkBox.isChecked())

                set_tab_by_name(self.ui.file_import_tabWidget, 'CSV')

                self.ui.import_csv_pushButton.setFocus()
                self.ui.import_excel_pushButton.clearFocus()

                # preview first N lines of input file
                self.ui.file_preview_tableWidget.setRowCount(0)
                self.ui.file_preview_tableWidget.setColumnCount(1)

                df = self.import_csv_file(preview=True, nrows=num_preview_rows)
                if df is not None and not self.ui.raw_file_preview_checkBox.isChecked():
                    self.preview_dataframe(df)
                else:
                    with open(file_pathname, 'r',
                              encoding=self.ui.import_csv_encoding_comboBox.currentText()) as f_read:
                        stop = False
                        if num_preview_rows:
                            for i in range(0, num_preview_rows):
                                if not stop:
                                    stop = self.try_readline(f_read, stop, i)
                        else:
                            i = 0
                            while not stop:
                                stop = self.try_readline(f_read, stop, i)
                                i += 1

                        self.ui.file_preview_tableWidget.horizontalHeader().setVisible(False)
                        self.ui.file_preview_tableWidget.resizeColumnsToContents()
                        self.ui.file_preview_tableWidget.horizontalHeader().setMinimumHeight(30)

    def preview_dataframe(self, df):
        # Set row and column count
        self.ui.file_preview_tableWidget.setRowCount(df.shape[0])
        self.ui.file_preview_tableWidget.setColumnCount(df.shape[1])

        # Set headers
        self.ui.file_preview_tableWidget.setHorizontalHeaderLabels(df.columns)

        # Populate the table
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                item = QTableWidgetItem(str(df.iloc[i, j]))
                self.ui.file_preview_tableWidget.setItem(i, j, item)

        self.ui.file_preview_tableWidget.horizontalHeader().setVisible(True)
        self.ui.file_preview_tableWidget.resizeColumnsToContents()
        self.ui.file_preview_tableWidget.horizontalHeader().setMinimumHeight(30)

    def handle_import_nans(self, df):
        if self.ui.column_drop_if_all_nans_radioButton.isChecked():
            df = df.dropna(axis=1, how='all')
        elif self.ui.column_drop_if_any_nans_radioButton.isChecked():
            df = df.dropna(axis=1, how='any')

        if self.ui.row_drop_if_all_nans_radioButton.isChecked():
            df = df.dropna(axis=0, how='all')
        elif self.ui.row_drop_if_any_nans_radioButton.isChecked():
            df = df.dropna(axis=0, how='any')

        return df

    def import_csv_file(self, preview=False, nrows=False, file_pathname=None):
        global status_bar_message, data, source_file_pathname

        if file_pathname is None:
            file_pathname = self.ui.filepathname_lineEdit.text()
            batch_mode = False
        else:
            batch_mode = True

        if file_pathname:
            source_file_pathname = file_pathname

            # TODO: reconnect these!
            # if not batch_mode:
            #     self.ui.export_data_lineEdit.setText(get_filename(source_file_pathname))
            #     self.init_on_import()

            delimiter = self.ui.import_csv_delimiter_comboBox.currentText()
            if delimiter == 'Auto':
                delimiter = None

            try:
                if self.ui.import_csv_skiprows_lineEdit.text():
                    skiprows = eval(self.ui.import_csv_skiprows_lineEdit.text())
                else:
                    skiprows = 0

                header_row = self.ui.import_csv_header_row_spinBox.value()

                if self.ui.import_csv_units_row_checkBox.isChecked():
                    units_row = self.ui.import_csv_units_row_spinBox.value()
                    skiprows += units_row - header_row
                else:
                    units_row = None

                try:
                    unitized_columns = get_unitized_columns(source_file_pathname,
                                                            encoding=self.ui.import_csv_encoding_comboBox.currentText(),
                                                            header_row=header_row, units_row=units_row,
                                                            delimiter=delimiter)

                    usecols = list(range(0, len(unitized_columns)))
                except:
                    unitized_columns = []
                    usecols = None

                keyword_args = self.import_csv_options_dict

                skip_blank_lines = eval(self.ui.import_csv_skip_blank_lines_comboBox.currentText())

                if preview:
                    try:
                        df = pd.read_csv(source_file_pathname, names=unitized_columns,
                                         header=header_row, delimiter=delimiter, usecols=usecols,
                                         encoding=self.ui.import_csv_encoding_comboBox.currentText(),
                                         skip_blank_lines=skip_blank_lines,
                                         nrows=nrows, skiprows=skiprows, on_bad_lines='warn',
                                         **keyword_args,
                                         )

                        df = self.handle_import_nans(df)
                    except:  # try without usecols...
                        try:
                            df = pd.read_csv(source_file_pathname, names=unitized_columns,
                                             header=header_row, delimiter=delimiter,  # usecols=usecols,
                                             encoding=self.ui.import_csv_encoding_comboBox.currentText(),
                                             skip_blank_lines=skip_blank_lines,
                                             nrows=nrows, skiprows=skiprows, on_bad_lines='warn',
                                             **keyword_args,
                                             )

                            df = self.handle_import_nans(df)
                        except:
                            df = None
                else:  # read in actual file
                    try:
                        df = pd.read_csv(source_file_pathname, names=unitized_columns, header=header_row,
                                         delimiter=delimiter, usecols=usecols,
                                         encoding=self.ui.import_csv_encoding_comboBox.currentText(),
                                         skip_blank_lines=skip_blank_lines, on_bad_lines='warn',
                                         skiprows=skiprows,
                                         **keyword_args,
                                         )

                        df = self.handle_import_nans(df)
                    except:  # try without usecols...
                        df = pd.read_csv(source_file_pathname, names=unitized_columns, header=header_row,
                                         delimiter=delimiter,  # usecols=usecols,
                                         encoding=self.ui.import_csv_encoding_comboBox.currentText(),
                                         skip_blank_lines=skip_blank_lines, on_bad_lines='warn',
                                         skiprows=skiprows,
                                         **keyword_args,
                                         )

                        df = self.handle_import_nans(df)

                    if not batch_mode:
                        data = df

                        self.setup_initial_triage_lists()

                        self.ui.statusbar.showMessage('imported %d rows of data, %d columns' %
                                                      (len(df), len(df.columns)), 10000)

                        self.ui.script_preview_consoleWidget.locals().update(globals())
                        self.ui.script_preview_consoleWidget.output.setPlainText('use run() to run user script!\n')

                return df

            except Exception as e:
                if not batch_mode:
                    QMessageBox(QMessageBox.Icon.Critical, 'CSV Import Error', 'Error reading "%s"\n\n%s' %
                                (file_pathname, str(e))).exec()
                else:
                    print('Excel Import Error')
                    print('Error reading "%s"\n\n%s' % (file_pathname, str(e)))

                return None

    def import_excel_file(self, preview=False, nrows=False, file_pathname=None):
        global status_bar_message, data, source_file_pathname

        if file_pathname is None:
            file_pathname = self.ui.filepathname_lineEdit.text()
            batch_mode = False
        else:
            batch_mode = True

        if file_pathname:
            source_file_pathname = file_pathname

            if not batch_mode:
                self.ui.export_data_lineEdit.setText(os.path.basename(file_pathname).rsplit('.', 1)[0])
                self.init_on_import()

            if self.ui.import_excel_sheet_comboBox.count() == 0:
                from openpyxl import load_workbook
                wb = load_workbook(filename=self.ui.filepathname_lineEdit.text(), read_only=True)
                self.ui.import_excel_sheet_comboBox.addItems(wb.sheetnames)
                self.ui.import_excel_sheet_comboBox.setCurrentIndex(0)

            sheet_name = self.ui.import_excel_sheet_comboBox.currentText()
            sheet_number = self.ui.import_excel_sheet_comboBox.currentIndex()

            header_row = self.ui.import_excel_header_row_spinBox.value()

            if self.ui.import_excel_units_row_checkBox.isChecked():
                units_row = self.ui.import_excel_units_row_spinBox.value()
            else:
                units_row = None

            try:
                if nrows is False:
                    nrows = None

                try:
                    # try loading by sheet name
                    unitized_columns = get_unitized_columns(source_file_pathname, sheet_name=sheet_name,
                                                            header_row=header_row, units_row=units_row)
                except:
                    # try loading by sheet number, in case sheet name changes across files...
                    unitized_columns = get_unitized_columns(source_file_pathname, sheet_name=sheet_number,
                                                            header_row=header_row, units_row=units_row)

                keyword_args = self.import_excel_options_dict

                engine_kwargs = {'read_only': True}

                if self.ui.import_excel_skiprows_lineEdit.text():
                    skiprows = eval(self.ui.import_excel_skiprows_lineEdit.text())
                else:
                    skiprows = None

                try:
                    # try loading by sheet name
                    df = pd.read_excel(source_file_pathname, names=unitized_columns, sheet_name=sheet_name,
                                       header=header_row, nrows=nrows, engine_kwargs=engine_kwargs, skiprows=skiprows,
                                       **keyword_args)
                except:
                    # try loading by sheet number, in case sheet name changes across files...
                    df = pd.read_excel(source_file_pathname, names=unitized_columns, sheet_name=sheet_number,
                                       header=header_row, nrows=nrows, engine_kwargs=engine_kwargs, skiprows=skiprows,
                                       **keyword_args)

                df = self.handle_import_nans(df)

                if not batch_mode:
                    self.ui.statusbar.showMessage('imported %d rows of df, %d columns' %
                                                  (len(df), len(df.columns)), 10000)

                    self.ui.script_preview_consoleWidget.locals().update(globals())
                    self.ui.script_preview_consoleWidget.output.setPlainText('use run() to run user script!\n')

                    if not preview:
                        data = df
                        self.setup_initial_triage_lists()

                return df

            except Exception as e:
                if not batch_mode:
                    QMessageBox(QMessageBox.Icon.Critical, 'Excel Import Error', 'Error reading "%s"\n\n%s' %
                                (file_pathname, str(e))).exec()
                else:
                    print('Excel Import Error')
                    print('Error reading "%s"\n\n%s' % (file_pathname, str(e)))
                return None

    @staticmethod
    def get_csv_help(self):
        doc_link = 'https://pandas.pydata.org/pandas-docs/version/%s/reference/api/pandas.read_csv.html' % pd.__version__

        if sys.platform.startswith('win'):
            os.system("start \"\" %s" % doc_link)
        else:
            os.system('open %s' % doc_link)

    @staticmethod
    def get_excel_help(self):
        doc_link = 'https://pandas.pydata.org/pandas-docs/version/%s/reference/api/pandas.read_excel.html' % pd.__version__

        if sys.platform.startswith('win'):
            os.system("start \"\" %s" % doc_link)
        else:
            os.system('open %s' % doc_link)

    def import_csv_header_row_changed(self):
        # set unit row to header row + 1, by default:
        if not self.ui.import_csv_units_row_checkBox.isChecked():
            self.ui.import_csv_units_row_spinBox.setValue(self.ui.import_csv_header_row_spinBox.value() + 1)
        self.load_file_preview()

    def import_excel_header_row_changed(self):
        if not self.ui.import_excel_units_row_checkBox.isChecked():
            # set unit row to header row + 1, by default:
            self.ui.import_excel_units_row_spinBox.setValue(self.ui.import_excel_header_row_spinBox.value() + 1)
        self.load_file_preview()

    def import_csv_freeform_changed(self):
        self.ui.import_csv_parameter_tableWidget.resizeColumnsToContents()

        self.import_csv_options_dict = (
            two_column_tableWidget_to_dict(self.ui.import_csv_parameter_tableWidget))

        self.load_file_preview()

    def import_excel_freeform_changed(self):
        self.ui.import_excel_parameter_tableWidget.resizeColumnsToContents()

        self.import_excel_options_dict = (
            two_column_tableWidget_to_dict(self.ui.import_excel_parameter_tableWidget))

        self.load_file_preview()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    widget = ImportTabWidget()
    widget.show()
    sys.exit(app.exec())
