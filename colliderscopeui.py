# This Python file uses the following encoding: utf-8
from __init__ import *

import os
import sys

path = os.path.dirname(os.path.abspath(__file__))
# sys.path.insert(0, os.path.join(path))  # picks up this package
sys.path.insert(0, os.path.join(path, path+os.sep+'assets'))  # picks up this package

os.chdir(path)
# print('colliderscopeui.py path = %s\n' % path)
# print('SYS Path = %s\n' % sys.path)
# print('CWD = %s\n' % os.getcwd())

import time
import pandas as pd
import numpy as np
import pyqtgraph as pg
from pyqtgraph import PlotWidget
from pyqtgraph.console import ConsoleWidget

import PySide6
from PySide6.QtCore import Qt, QObject, QThread, Signal
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QDialog, QFileDialog, QTableWidget,
                               QTableWidgetItem, QLabel, QMessageBox, QGraphicsScene, QGraphicsWidget,
                               QGraphicsProxyWidget, QSizePolicy, QVBoxLayout, QHBoxLayout, QCheckBox,)

from PySide6.QtGui import QStandardItemModel, QBrush, QColor, QTextCursor, QFont

from pythonhighlighter import PythonHighlighter

# Important:
# You need to run the following command to generate the ui_form.py file or Build/Run in Qt Creator first:
#     pyside6-uic form.ui -o ui_form.py, or
#     pyside2-uic form.ui -o ui_form.py
from ui_colliderscope import Ui_ColliderScopeUI
from ui_nanhandler_horizontal import Ui_NanHandlerHorizontal

from file_io import *

timer = None

app = None
mainwindow = None
data = None
status_bar_message = ''
latest_item = None
latest_items = None

source_file_pathname = ''

original_numeric_fields = []
original_string_fields = []

active_numeric_fields = []
active_string_fields = []
ignore_fields = []
favorite_fields = []

import_nan_handling = None
export_nan_handling = None

pg.setConfigOptions(antialias=False)
# pg.setConfigOption('background', 'w')
# pg.setConfigOption('foreground', 'b')

# from ui_nan_handler_dialog import Ui_NanHandlerDialog


def file_dialog(file_pathname, file_type_filter, file_dialog_title):
    """
    Opens a file dialog to select a file with extension options.

    :param file_pathname: Default file name
    :param file_type_filter: Specifies extension filter type
    :param file_dialog_title: Title for dialog box
    :return: User selected file name, Echo file_type, Echo file_dialog_title
    """

    dialog = QFileDialog()
    dialog.selectFile(file_pathname)
    dialog.setWindowTitle(file_dialog_title)
    dialog.setFileMode(QFileDialog.ExistingFile)
    # dialog.setFileMode(QFileDialog.AnyFile)
    if type(file_type_filter) is not list:
        dialog.setNameFilters([file_type_filter])
    else:
        dialog.setNameFilters(file_type_filter)
    dialog.setViewMode(QFileDialog.Detail)
    if dialog.exec():
        file_pathname = dialog.selectedFiles()
        file_pathname = str(file_pathname)[2:-2]
        return file_pathname
    else:
        file_pathname = ""
        return file_pathname


def nan_handler_dialog():
    dialog = QDialog()
    nan_dialog = Ui_NanHandlerDialog()
    nan_dialog.setupUi(dialog)
    dialog.exec()

    nan_handling = {0: dict(), 1: dict()}  # 0 = rows, 1 = columns

    nan_handling[0]['drop_if_all_nans'] = nan_dialog.row_drop_if_all_nans_radioButton.isChecked()
    nan_handling[0]['drop_if_any_nans'] = nan_dialog.row_drop_if_any_nans_radioButton.isChecked()
    nan_handling[1]['drop_if_all_nans'] = nan_dialog.column_drop_if_all_nans_radioButton.isChecked()
    nan_handling[1]['drop_if_any_nans'] = nan_dialog.column_drop_if_any_nans_radioButton.isChecked()

    return nan_handling


def two_column_tableWidget_to_dict(table_widget):
    """
    Convert a QTableWidget (instance of a table widget in PyQt) to a dictionary using data from its first
    two columns.

    Parameters:
        table_widget (QTableWidget): The instance of the table widget.

    Returns:
        dict: A dictionary where keys and values are textual data from the first and second
              columns of each row in the table_widget. If a row doesn't have items in both columns,
              it will be ignored.

    """
    options_dict = dict()

    for r in range(table_widget.rowCount()):
        parameter = table_widget.item(r, 0)
        parameter_value = table_widget.item(r, 1)

        if (parameter is not None and
                parameter.text() != '' and
                parameter_value is not None and
                parameter_value.text() != ''):
            if parameter_value.text().lower() == 'none':
                options_dict[parameter.text()] = None
            elif parameter_value.text() == "''":
                options_dict[parameter.text()] = ''
            elif parameter_value.text().lower() == 'false':
                options_dict[parameter.text()] = False
            elif parameter_value.text().lower() == 'true':
                options_dict[parameter.text()] = True
            else:
                try:
                    options_dict[parameter.text()] = eval(parameter_value.text())
                except:
                    options_dict[parameter.text()] = parameter_value.text()

    return options_dict


def get_unitized_columns(filename, sheet_name=None, ignore_units=(), encoding='utf-8', header_row=0, units_row=None,
                         skiprows=None, delimiter=None):
    """
    Combine column labels and units row into a single combined string to identify the column.

    Args:
        filename (str): name of the file to read
        sheet_name (str): for reading a particular Excel sheet, or ``None`` for CSV files
        ignore_units (list of str): unit values to ignore, default ``Text``
        encoding (str): file encoding (decoding) method name
        header_row (int): the header row number
        units_row (int): the unit row number, if any

    Returns:
        List of combined column headers and units as in ``['ColumnName_units', ...]``

    """
    num_rows = 100  # may have to read a chunk of row2 if there are partial columns...??

    if sheet_name:
        columns = pd.read_excel(filename, header=None, nrows=num_rows, skiprows=list(range(0, header_row)),
                                sheet_name=sheet_name)

        if units_row is not None:
            units = pd.read_excel(filename, header=None, nrows=num_rows, skiprows=list(range(0, units_row)),
                                  sheet_name=sheet_name)
        else:
            units = pd.DataFrame({'units': [''] * columns.shape[1]}).transpose()
    else:
        columns = pd.read_csv(filename, header=None, nrows=num_rows, skiprows=list(range(0, header_row)),
                              encoding=encoding, encoding_errors='strict', on_bad_lines='skip', delimiter=delimiter)

        if units_row is not None:
            units = pd.read_csv(filename, header=None, nrows=num_rows, skiprows=list(range(0, units_row)),
                                encoding=encoding, encoding_errors='strict', on_bad_lines='skip', delimiter=delimiter)

            # pad units with blanks if fewer units than columns...
            for i in range(max(0, len(columns.values[0]) - len(units.values[0]))):
                units[' ' * i] = ''

        else:  # blank units
            units = pd.DataFrame({'units': [''] * columns.shape[1]}).transpose()

    unitized_columns = []

    for col, unit in zip(columns.values[0], units.values[0]):
        if unit not in ignore_units and pd.notna(unit) and units_row is not None:
            if str(unit).strip() != '':
                unitized_columns.append('%s_%s' % (str(col).strip(), str(unit).strip()))
            else:
                unitized_columns.append('%s' % str(col).strip())
        else:
            unitized_columns.append('%s' % str(col).strip())

    # deal with dupicates!
    sanitized_columns = []
    raw_columns = []

    for col in unitized_columns:
        raw_columns.append(col)
        if col not in sanitized_columns:
            sanitized_columns.append(col)
        else:
            sanitized_columns.append('%s:%d' % (col, raw_columns.count(col)))

    return sanitized_columns


def mouseMoved(*args, **kwargs):
    print('mouseMoved', args)


def mouseHover(*args, **kwargs):
    print('mouseHover', args)


def mouseClicked(*args, **kwargs):
    print('mouseClicked', args)


def run():
    exec(mainwindow.ui.script_preview_plainTextEdit.toPlainText(), globals())


def ignore_listWidget_dropEvent(event):
    item_list = [i.text() for i in event.source().selectedItems()]

    for i in item_list:
        event.source().source_data.remove(i)
        mainwindow.ui.triage_ignore_listWidget.source_data.append(i)

    mainwindow.ui.triage_ignore_listWidget.addItems(item_list)

    event.setDropAction(PySide6.QtCore.Qt.DropAction.MoveAction)
    event.accept()

    mainwindow.ui.triage_filter_widget.inputChanged()  # update triage listWidgets


def favorites_listWidget_dropEvent(event):
    item_list = [i.text() for i in event.source().selectedItems()]

    for i in item_list:
        event.source().source_data.remove(i)
        mainwindow.ui.triage_favorites_listWidget.source_data.append(i)

    mainwindow.ui.triage_favorites_listWidget.addItems(item_list)

    event.setDropAction(PySide6.QtCore.Qt.DropAction.MoveAction)
    event.accept()


def set_tab_by_name(tabWidget, tab_name):
    for i in range(tabWidget.count()):
        if tabWidget.tabText(i) == tab_name:
            tabWidget.setCurrentIndex(i)


def get_current_tabname(tabWidget):
    return tabWidget.tabText(tabWidget.currentIndex())


def try_to_float(v):
    try:
        return float(v)
    except:
        return np.nan


class ExportWorker(QObject):
    finished = Signal()
    progress = Signal(int)

    def __init__(self, source_files):
        super().__init__()
        self.source_files = source_files
        self.cancelled = False

    def handle_export_nans(self, df):
        if mainwindow.ui.nanhandler_widget.ui.column_drop_if_all_nans_radioButton.isChecked():
            df = df.dropna(axis=1, how='all')
        elif mainwindow.ui.nanhandler_widget.ui.column_drop_if_any_nans_radioButton.isChecked():
            df = df.dropna(axis=1, how='any')

        if mainwindow.ui.nanhandler_widget.ui.row_drop_if_all_nans_radioButton.isChecked():
            df = df.dropna(axis=0, how='all')
        elif mainwindow.ui.nanhandler_widget.ui.row_drop_if_any_nans_radioButton.isChecked():
            df = df.dropna(axis=0, how='any')

        return df

    def export_files(self):
        global data

        combined_dfs = []
        source_filenames = []
        skipped_filenames = []  # TODO: add this as an attribute and handle skip messaging in main thread

        for idx, source_file in enumerate(self.source_files):
            if not self.cancelled:
                file_name = get_filename(source_file)
                source_filenames.append(file_name)

                if get_current_tabname(mainwindow.ui.file_import_tabWidget) == 'CSV':
                    data = mainwindow.import_csv_file(file_pathname=source_file)
                else:
                    data = mainwindow.import_excel_file(file_pathname=source_file)

                if data is not None:
                    # run the script for calculated values, etc, but don't update triage lists, etc
                    mainwindow.ui.script_preview_consoleWidget.repl.runCmd('run()')

                    data = self.handle_export_nans(data)

                    if 'Combined' in mainwindow.ui.export_data_mode_comboBox.currentText():
                        combined_dfs.append(data.copy())

                    if mainwindow.ui.export_data_mode_comboBox.currentText() != 'Combined':
                        # export source file(s) unless mode is 'Combined' only
                        mainwindow.export_file(data, file_name)
                else:
                    skipped_filenames.append(source_file)
                    print('!!! Skipping %s, no data loaded !!!' % source_file)

                mainwindow.ui.export_progressBar.setMaximum(len(self.source_files))

            self.progress.emit(idx)

        if 'Combined' in mainwindow.ui.export_data_mode_comboBox.currentText():
            combined_filename = create_combined_filename(source_filenames)

            combined_df = pd.concat(combined_dfs, ignore_index=True, sort=False)

            combined_df = self.handle_export_nans(combined_df)

            mainwindow.export_file(combined_df, combined_filename + '-combined')

        # if skipped_filenames:
        #     sf_str = '\n'
        #     for sf in skipped_filenames:
        #         sf_str += 'sf\n'
        #
        #     QMessageBox(QMessageBox.Icon.Warning,
        #                 'Export Warning', "Unable to export %d files: %s" % (len(skipped_filenames), '')).exec()

        self.finished.emit()

    def cancel(self):
        self.cancelled = True


def preprocess_script_keyPressEvent(event):
    if event.key() == Qt.Key_Home:
        cursor = mainwindow.ui.script_preview_plainTextEdit.textCursor()
        cursor.movePosition(QTextCursor.StartOfLine)
        mainwindow.ui.script_preview_plainTextEdit.setTextCursor(cursor)
    elif event.key() == Qt.Key_End:
        cursor = mainwindow.ui.script_preview_plainTextEdit.textCursor()
        cursor.movePosition(QTextCursor.EndOfLine)
        mainwindow.ui.script_preview_plainTextEdit.setTextCursor(cursor)
    else:
        mainwindow.ui.script_preview_plainTextEdit.default_keyPressEvent(event)


def import_csv_header_row_spinBox_stepBy(step):
    mainwindow.ui.import_csv_header_row_spinBox.default_stepBy(step)
    mainwindow.load_file_preview()


def import_csv_units_row_spinBox_stepBy(step):
    mainwindow.ui.import_csv_units_row_spinBox.default_stepBy(step)
    mainwindow.load_file_preview()


def import_excel_header_row_spinBox_stepBy(step):
    mainwindow.ui.import_excel_header_row_spinBox.default_stepBy(step)
    mainwindow.load_file_preview()


def import_excel_units_row_spinBox_stepBy(step):
    mainwindow.ui.import_excel_units_row_spinBox.default_stepBy(step)
    mainwindow.load_file_preview()


def preview_size_spinBox_stepBy(step):
    mainwindow.ui.preview_size_spinBox.default_stepBy(step)
    mainwindow.load_file_preview()


class ColliderScopeUI(QMainWindow):
    def __init__(self, parent=None):
        global active_numeric_fields, active_string_fields, ignore_fields, favorite_fields

        super().__init__(parent)
        self.ui = Ui_ColliderScopeUI()
        self.ui.setupUi(self)

        # make sure we start on the import tab in case the .ui was left on another tab
        set_tab_by_name(self.ui.tabWidget_main, 'Import')

        self.init_plot_widget(self.ui.graphic_preview_plot_widget, 'Graphic Preview')

        self.ui.file_import_browse_pushButton.setFocus()

        self.import_csv_options_dict = dict()
        self.import_excel_options_dict = dict()
        self.ui.import_csv_parameter_tableWidget.horizontalHeader().setMinimumHeight(25)
        self.ui.import_excel_parameter_tableWidget.horizontalHeader().setMinimumHeight(25)

        self.hl = PythonHighlighter(self.ui.script_preview_plainTextEdit.document())

        self.ui.triage_numeric_listWidget.source_data = active_numeric_fields
        self.ui.triage_numeric_listWidget.count_label = self.ui.active_numeric_count_label
        self.ui.triage_numeric_listWidget.active_label = None

        self.ui.triage_string_listWidget.source_data = active_string_fields
        self.ui.triage_string_listWidget.count_label = self.ui.active_string_count_label
        self.ui.triage_string_listWidget.active_label = None

        self.ui.triage_ignore_listWidget.source_data = ignore_fields
        self.ui.triage_ignore_listWidget.count_label = self.ui.ignore_count_label
        self.ui.triage_ignore_listWidget.active_label = None

        self.ui.triage_favorites_listWidget.source_data = favorite_fields
        self.ui.triage_favorites_listWidget.count_label = None
        self.ui.triage_favorites_listWidget.active_label = None

        self.ui.text_preview_listWidget.count_label = None
        self.ui.text_preview_listWidget.active_label = self.ui.text_preview_active_count_label

        self.triage_listwidgets = [
            self.ui.triage_numeric_listWidget,
            self.ui.triage_string_listWidget,
            self.ui.triage_ignore_listWidget,
            self.ui.triage_favorites_listWidget
        ]

        self.active_listwidget = None

        self.ui.triage_filter_widget.widget_list = \
            [self.ui.triage_numeric_listWidget, self.ui.triage_string_listWidget,
             self.ui.triage_ignore_listWidget, self.ui.triage_favorites_listWidget]

        self.ui.text_preview_filter_widget.widget_list = [self.ui.text_preview_listWidget]

        self.ui.triage_ignore_listWidget.dropEvent = ignore_listWidget_dropEvent
        self.ui.triage_favorites_listWidget.dropEvent = favorites_listWidget_dropEvent

        self.prior_nan_count = None

        # this is how to instantiate a re-usable widget created in Qt Creator via a .ui file,
        # do NOT "promote" in the Qt Creator, it won't run, the setupUi() will fail since there's no way to pass in
        # the target widget in the compiled python script. The filter widget has a proper class, and the ui part
        # gets handled in the __init__, but the nanhandler doesn't have a proper class, just the ui.  The filter
        # widget was an entire Qt Creator "project", not just an isolated ui file.
        self.ui.nanhandler_widget.ui = Ui_NanHandlerHorizontal()
        self.ui.nanhandler_widget.ui.setupUi(self.ui.nanhandler_widget)

        self.active_plot = pg.plot()
        self.init_plot_widget(self.active_plot, '')

        # # attempt at making two independent grids line up...
        # min_grid_widths = [41, 100, 100, 100, 41, 100, 41, 100, 41]
        # stretch = [0, 1, 1, 1, 0, 1, 0, 1, 0]
        #
        # for idx, gw in enumerate(min_grid_widths):
        #     self.ui.plot_inputs_label_gridLayout.setColumnMinimumWidth(idx, gw)
        #     self.ui.plot_inputs_gridLayout.setColumnMinimumWidth(idx, gw)
        #
        # for idx, s in enumerate(stretch):
        #     self.ui.plot_inputs_label_gridLayout.setColumnStretch(idx, s)
        #     self.ui.plot_inputs_gridLayout.setColumnStretch(idx, s)

        self.ui.scrollAreaWidgetContents_2.setVisible(False)

        # import/export-related attributes
        self.input_file_pathname = ''

        self.export_folder_pathname = ''
        self.export_worker_thread = None
        self.export_worker = None
        self.export_batch_source_files = []
        self.pre_export_data = None

        # Connect keyPressEvent to handle keyboard events in export batch list widget
        self.ui.export_batch_files_listWidget.default_keyPressEvent = (
            self.ui.export_batch_files_listWidget.keyPressEvent)
        self.ui.export_batch_files_listWidget.keyPressEvent = self.remove_export_batch_files

        # set up drag/drop events for export_batch_files_listWidget
        self.ui.export_batch_files_listWidget.dragEnterEvent = self.export_batch_files_listWidget_dragEnterEvent
        self.ui.export_batch_files_listWidget.dropEvent = self.export_batch_files_listWidget_dropEvent
        #   dragMoveEvent MUST be set up or QListWidget will not receive drops!
        self.ui.export_batch_files_listWidget.dragMoveEvent = self.export_batch_files_listWidget_dragMoveEvent

        # handle preprocess script keyPressEvents for Home and End keys
        self.ui.script_preview_plainTextEdit.default_keyPressEvent = \
            self.ui.script_preview_plainTextEdit.keyPressEvent
        self.ui.script_preview_plainTextEdit.keyPressEvent = preprocess_script_keyPressEvent

        # intercept import spinBox button step events to trigger preview update without using valueChanged() signal
        self.ui.import_csv_header_row_spinBox.default_stepBy = self.ui.import_csv_header_row_spinBox.stepBy
        self.ui.import_csv_header_row_spinBox.stepBy = import_csv_header_row_spinBox_stepBy

        self.ui.import_csv_units_row_spinBox.default_stepBy = self.ui.import_csv_units_row_spinBox.stepBy
        self.ui.import_csv_units_row_spinBox.stepBy = import_csv_units_row_spinBox_stepBy

        self.ui.import_excel_header_row_spinBox.default_stepBy = self.ui.import_excel_header_row_spinBox.stepBy
        self.ui.import_excel_header_row_spinBox.stepBy = import_excel_header_row_spinBox_stepBy

        self.ui.import_excel_units_row_spinBox.default_stepBy = self.ui.import_excel_units_row_spinBox.stepBy
        self.ui.import_excel_units_row_spinBox.stepBy = import_excel_units_row_spinBox_stepBy

        self.ui.preview_size_spinBox.default_stepBy = self.ui.preview_size_spinBox.stepBy
        self.ui.preview_size_spinBox.stepBy = preview_size_spinBox_stepBy

        # customize font in the python interpreter input text edit
        font = QFont()
        font.setFamily("Courier New")
        font.setStyleStrategy(QFont.StyleStrategy.PreferAntialias)
        self.ui.script_preview_consoleWidget.repl.input.setFont(font)

        # timer.start()

        pass

    def generic_slot(self, *args, **kwargs):
        print('generic slot...')
        print(self.ui.triage_favorites_listWidget.selectedItems())

    def send_right_pushbutton(self):
        if self.ui.ignore_favorites_tabWidget.currentIndex() == 0:
            destination = self.ui.triage_ignore_listWidget
        else:
            destination = self.ui.triage_favorites_listWidget

        for lw in [self.ui.triage_numeric_listWidget, self.ui.triage_string_listWidget]:
            selected_fields = [s.text() for s in lw.selectedItems()]
            for sf in selected_fields:
                lw.source_data.remove(sf)
                destination.source_data.append(sf)

        self.ui.triage_filter_widget.inputChanged()  # update triage listWidgets

    def send_left_pushbutton(self):
        if self.ui.ignore_favorites_tabWidget.currentIndex() == 0:
            source = self.ui.triage_ignore_listWidget
        else:
            source = self.ui.triage_favorites_listWidget

        selected_fields = [s.text() for s in source.selectedItems()]
        for sf in selected_fields:
            source.source_data.remove(sf)
            if sf in original_numeric_fields:
                active_numeric_fields.append(sf)
            else:
                active_string_fields.append(sf)

        self.ui.triage_filter_widget.inputChanged()  # update triage listWidgets

    @staticmethod
    def get_csv_help():
        doc_link = 'https://pandas.pydata.org/pandas-docs/version/%s/reference/api/pandas.read_csv.html' % pd.__version__

        if sys.platform.startswith('win'):
            os.system("start \"\" %s" % doc_link)
        else:
            os.system('open %s' % doc_link)

    @staticmethod
    def get_excel_help():
        doc_link = 'https://pandas.pydata.org/pandas-docs/version/%s/reference/api/pandas.read_excel.html' % pd.__version__

        if sys.platform.startswith('win'):
            os.system("start \"\" %s" % doc_link)
        else:
            os.system('open %s' % doc_link)

    def init_plot_widget(self, plot_widget, title_str):
        plot_widget.plotItem.clear()
        plot_widget.showGrid(x=True, y=True)
        plot_widget.plotItem.setTitle(title_str)
        plot_widget.plotItem.showButtons()
        plot_widget.new = True

        # set up selection rectangle
        plot_widget.getViewBox().setMouseMode(pg.ViewBox.RectMode)
        plot_widget.getViewBox().rbScaleBox.setPen(pg.mkPen((64, 128, 200), width=2))
        plot_widget.getViewBox().rbScaleBox.setBrush(pg.mkBrush(81,	197, 255, 100))

        # plot_widget.plotItem.ctrlMenu # for future work...

        # plot_widget.scene().sigMouseMoved.connect(mouseMoved)
        # plot_widget.scene().sigMouseHover.connect(mouseHover)
        plot_widget.scene().sigMouseClicked.connect(mouseClicked)

        backgroundBrush = QBrush(QColor(7, 27, 46, 255))
        backgroundBrush.setStyle(Qt.SolidPattern)
        plot_widget.setBackgroundBrush(backgroundBrush)

        # self.ui.plot_graphicsView.showGrid(x=True, y=True)

    def try_readline(self, f_read, stop, i):
        try:
            line = f_read.readline()
        except Exception as e:
            line = str(e)
            stop = True

        if line:
            self.ui.file_preview_tableWidget.insertRow(self.ui.file_preview_tableWidget.rowCount())
            # label rows from 0 and not 1, to be compatible with import settings like header_row:
            self.ui.file_preview_tableWidget.setVerticalHeaderItem(i, QTableWidgetItem(str(i)))
            self.ui.file_preview_tableWidget.setItem(i-1, 1, QTableWidgetItem('%s' % line.rstrip()))
        else:
            stop = True

        return stop

    def load_file_preview(self, qstring='', file_pathname=None):
        print('load_file_preview')

        self.ui.file_import_browse_pushButton.clearFocus()

        if self.ui.preview_size_checkBox.isChecked():
            self.ui.preview_size_spinBox.setEnabled(True)
            num_preview_rows = self.ui.preview_size_spinBox.value()
        else:
            self.ui.preview_size_spinBox.setEnabled(False)
            num_preview_rows = None

        if file_pathname is None:
            file_pathname = self.ui.filepathname_lineEdit.text()

        if file_pathname:
            if 'xls' in file_pathname.split('.')[-1]:
                self.ui.import_excel_units_row_spinBox.setEnabled(self.ui.import_excel_units_row_checkBox.isChecked())

                set_tab_by_name(self.ui.file_import_tabWidget, 'Excel')

                self.ui.import_excel_pushButton.setFocus()
                self.ui.import_csv_pushButton.clearFocus()

                df = self.import_excel_file(preview=True, nrows=num_preview_rows)
                if df is not None:
                    self.preview_dataframe(df)

            else:
                self.ui.import_csv_units_row_spinBox.setEnabled(self.ui.import_csv_units_row_checkBox.isChecked())

                set_tab_by_name(self.ui.file_import_tabWidget, 'CSV')

                self.ui.import_csv_pushButton.setFocus()
                self.ui.import_excel_pushButton.clearFocus()

                # preview first N lines of input file
                self.ui.file_preview_tableWidget.setRowCount(0)
                self.ui.file_preview_tableWidget.setColumnCount(1)

                df = self.import_csv_file(preview=True, nrows=num_preview_rows)
                if df is not None and not self.ui.raw_file_preview_checkBox.isChecked():
                    self.preview_dataframe(df)
                else:
                    with open(file_pathname, 'r',
                              encoding=self.ui.import_csv_encoding_comboBox.currentText()) as f_read:
                        stop = False
                        if num_preview_rows:
                            for i in range(0, num_preview_rows):
                                if not stop:
                                    stop = self.try_readline(f_read, stop, i)
                        else:
                            i = 0
                            while not stop:
                                stop = self.try_readline(f_read, stop, i)
                                i += 1

                        self.ui.file_preview_tableWidget.horizontalHeader().setVisible(False)
                        self.ui.file_preview_tableWidget.resizeColumnsToContents()
                        self.ui.file_preview_tableWidget.horizontalHeader().setMinimumHeight(30)

    def import_csv_header_row_changed(self):
        # set unit row to header row + 1, by default:
        if not self.ui.import_csv_units_row_checkBox.isChecked():
            self.ui.import_csv_units_row_spinBox.setValue(self.ui.import_csv_header_row_spinBox.value() + 1)
        self.load_file_preview()

    def import_excel_header_row_changed(self):
        if not self.ui.import_excel_units_row_checkBox.isChecked():
            # set unit row to header row + 1, by default:
            self.ui.import_excel_units_row_spinBox.setValue(self.ui.import_excel_header_row_spinBox.value() + 1)
        self.load_file_preview()

    def preview_dataframe(self, df):
        # Set row and column count
        self.ui.file_preview_tableWidget.setRowCount(df.shape[0])
        self.ui.file_preview_tableWidget.setColumnCount(df.shape[1])

        # Set headers
        self.ui.file_preview_tableWidget.setHorizontalHeaderLabels(df.columns)

        # Populate the table
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                item = QTableWidgetItem(str(df.iloc[i, j]))
                self.ui.file_preview_tableWidget.setItem(i, j, item)

        self.ui.file_preview_tableWidget.horizontalHeader().setVisible(True)
        self.ui.file_preview_tableWidget.resizeColumnsToContents()
        self.ui.file_preview_tableWidget.horizontalHeader().setMinimumHeight(30)

    def handle_import_button_enables(self):
        if self.ui.filepathname_lineEdit.text():
            self.ui.file_import_tabWidget.setEnabled(True)
            self.ui.row_groupBox.setEnabled(True)
            self.ui.column_groupBox.setEnabled(True)
        else:
            self.ui.file_import_tabWidget.setEnabled(False)
            self.ui.row_groupBox.setEnabled(False)
            self.ui.column_groupBox.setEnabled(False)

    def filepathname_changed(self):
        self.load_file_preview('', self.ui.filepathname_lineEdit.text())
        self.handle_import_button_enables()

    def import_filebrowse(self):
        self.input_file_pathname = file_dialog('', '*.*', 'file_dialog_title')
        self.ui.filepathname_lineEdit.setText(self.input_file_pathname)
        self.ui.import_excel_sheet_comboBox.clear()
        self.handle_import_button_enables()

        self.load_file_preview('', self.input_file_pathname)

    def numeric_listwidget_selection_changed(self):
        self.active_listwidget = self.ui.triage_numeric_listWidget
        self.handle_active_listwidget()

    def string_listwidget_selection_changed(self):
        self.active_listwidget = self.ui.triage_string_listWidget
        self.handle_active_listwidget()

    def ignore_listwidget_selection_changed(self):
        self.active_listwidget = self.ui.triage_ignore_listWidget
        self.handle_active_listwidget()

    def favorites_listwidget_selection_changed(self):
        self.active_listwidget = self.ui.triage_favorites_listWidget
        self.handle_active_listwidget()

    def handle_active_listwidget(self, force_native_preview=False):
        global latest_item, latest_items

        for lw in self.triage_listwidgets:
            if lw is not self.active_listwidget:
                lw.clearSelection()

        if self.active_listwidget.currentItem() is not None:
            latest_item = self.active_listwidget.currentItem().text()
        latest_items = [i.text() for i in self.active_listwidget.selectedItems()]

        if latest_item in original_string_fields:
            self.update_string_preview(force_native_preview)
        else:
            self.update_numeric_preview(force_native_preview)

    def triage_doubleclick_selection_handler(self, current_listWidgetItem):
        self.active_listwidget = current_listWidgetItem.listWidget()
        self.handle_active_listwidget(force_native_preview=True)

    def update_text_preview(self, latest_item):
        self.ui.text_preview_listWidget.clear()
        self.ui.text_preview_listWidget.source_data = []

        if latest_item in data:
            self.ui.text_preview_listWidget.addItems(['%s:' % latest_item, '-' * (len(latest_item) + 1)])

            unique_values = data[latest_item].unique()
            self.ui.text_preview_unique_checkBox.setText('Unique [%d]' % len(unique_values))

            if self.ui.text_preview_unique_checkBox.isChecked():
                self.ui.text_preview_listWidget.source_data = [str(d) for d in unique_values]
            else:
                self.ui.text_preview_listWidget.source_data = [str(d) for d in data[latest_item]]

            self.ui.text_preview_listWidget.addItems(self.ui.text_preview_listWidget.source_data)

            self.ui.text_preview_filter_widget.inputChanged()

    def update_string_preview(self, force=False):
        if get_current_tabname(self.ui.preview_tabWidget) != 'Preprocess Script' or force:
            set_tab_by_name(self.ui.preview_tabWidget, 'Text Preview')

        self.update_text_preview(latest_item)

    def update_numeric_preview(self, force=False):
        if force:
            set_tab_by_name(self.ui.preview_tabWidget, 'Graphic Preview')

        self.update_text_preview(latest_item)

        if latest_item in data:
            nan_count = sum(data[latest_item].isna())
            allow_autorange = True

            if (self.ui.graphic_preview_plot_widget.new or
                    nan_count != self.prior_nan_count or
                    sum(data[latest_item].notna()) == 0):
                # self.ui.graphic_preview_plot_widget.plot(data[latest_item].values, pen=None,
                #                                      symbolBrush=(231, 232, 255), symbolPen=(231, 232, 255), symbol='o',
                #                                      symbolSize=1.5, clear=True)
                # self.ui.graphic_preview_plot_widget.plot(data[latest_item].values, pen=None,
                #                                      symbolBrush=None, symbolPen=(231, 232, 255), symbol='t1',
                #                                      symbolSize=4, clear=True)

                # self.ui.graphic_preview_plot_widget.plot(data[latest_item].values, pen=(231, 232, 255), clear=True)

                if sum(data[latest_item].isna()) > 0 or len(data[latest_item]) == 1:
                    if sum(data[latest_item].notna()) == 0:
                        # data is all nans...
                        self.ui.graphic_preview_plot_widget.plot([0], [0], clear=True)
                        self.ui.graphic_preview_plot_widget.setXRange(-0.5, 0.5)
                        self.ui.graphic_preview_plot_widget.setYRange(-0.5, 0.5)
                        allow_autorange = False
                    else:
                        self.ui.graphic_preview_plot_widget.plot(data.index, data[latest_item].values, pen='#00000000',
                                                             symbolBrush=None, symbolPen=(231, 232, 255),
                                                             symbol='t1', symbolSize=4, clear=True)
                        pts = data[latest_item].isna()
                        self.ui.graphic_preview_plot_widget.plot(data[latest_item].loc[pts].index,
                                                                 [0] * len(data[latest_item].loc[pts].index),
                                                                 pen='#00000000', symbolBrush=(255, 80, 80, 100),
                                                                 symbolPen=(255, 80, 80, 100), symbol='o', symbolSize=8)

                    # label nan count, "anchor" is relative to upper left corner of the box
                    text = pg.TextItem(
                        html='<div style="text-align: center"><span style="color: #FFF;'
                             '"<span style="color: #FF0; font-size: 16pt;">%d NaNs</span></div>' % nan_count,
                        anchor=(-0.03*0, 0.5*0), border='w', fill=(255, 80, 80, 100))
                    self.ui.graphic_preview_plot_widget.addItem(text)

                    if sum(data[latest_item].notna()) == 0:
                        # data is all nans...
                        text.setPos(0, 0)
                    else:
                        text.setPos(0, data[latest_item].max())
                else:
                    self.ui.graphic_preview_plot_widget.plot(data.index, data[latest_item].values, pen=(231, 232, 255),
                                                             clear=True)

                self.ui.graphic_preview_plot_widget.new = False
            else:  # re-using plot and no nans, just update data
                self.ui.graphic_preview_plot_widget.plotItem.curves[0].setData(data[latest_item].values)

            self.ui.graphic_preview_plot_widget.plotItem.setTitle(latest_item)
            if allow_autorange:
                self.ui.graphic_preview_plot_widget.plotItem.autoRange()
            # maybe do this if len(data) > X?
            # self.ui.graphic_preview_plot_widget.setDownsampling(auto=True, mode='peak')

            self.prior_nan_count = nan_count
        else:
            self.prior_nan_count = 0
            self.init_plot_widget(self.ui.graphic_preview_plot_widget, 'Graphic Preview')

    def init_triage_lists(self):
        global data
        data = data.convert_dtypes(convert_boolean=False)
        active_numeric_fields.clear()
        active_string_fields.clear()
        ignore_fields.clear()
        favorite_fields.clear()

        drop_fields = set()
        # convert mixed columns (e.g. strings and numbers) to strings
        non_numeric_fields = [c for c in data.select_dtypes(exclude=['number']).columns]
        for c in non_numeric_fields:
            if not pd.api.types.is_string_dtype(data[c]):
                data['%s__obj->str' % c] = data[c].astype('string')
                drop_fields.add(c)
                try:
                    data['%s__obj->float' % c] = data[c].as_type('float')
                    drop_fields.add(c)
                except:
                    data['%s__obj->float' % c] = data[c].apply(try_to_float)
                    drop_fields.add(c)

        data = data.drop(drop_fields, axis=1)

        active_numeric_fields.extend(data.select_dtypes(include='number').columns)
        active_string_fields.extend(data.select_dtypes(include='string').columns)

        self.ui.triage_filter_widget.inputChanged()  # update triage widgets

    def ignore_deadvars(self):
        global ignore_fields

        for c in data.columns:
            if sum(data[c].notna()) == 0:
                ignore_fields.append(c)
            else:
                try:
                    if data[c].min() == data[c].max():
                        ignore_fields.append(c)
                except:
                    print(c)

        ignore_fields = list(np.unique(ignore_fields))
        self.ui.triage_ignore_listWidget.source_data = ignore_fields

        for ignore_field in ignore_fields:
            if ignore_field in active_numeric_fields:
                active_numeric_fields.remove(ignore_field)
            if ignore_field in active_string_fields:
                active_string_fields.remove(ignore_field)

        self.ui.triage_filter_widget.inputChanged()  # update triage widgets

        set_tab_by_name(self.ui.ignore_favorites_tabWidget, 'Ignore')

    def handle_import_nans(self, df):
        if self.ui.column_drop_if_all_nans_radioButton.isChecked():
            df = df.dropna(axis=1, how='all')
        elif self.ui.column_drop_if_any_nans_radioButton.isChecked():
            df = df.dropna(axis=1, how='any')

        if self.ui.row_drop_if_all_nans_radioButton.isChecked():
            df = df.dropna(axis=0, how='all')
        elif self.ui.row_drop_if_any_nans_radioButton.isChecked():
            df = df.dropna(axis=0, how='any')

        return df

    def setup_initial_triage_lists(self):
        global data
        self.ui.tabWidget_main.setCurrentIndex(1)
        self.init_triage_lists()
        # grab non-filtered original fields for later reference:
        original_numeric_fields.extend(active_numeric_fields)
        original_string_fields.extend(active_string_fields)

        self.ui.triage_tab.setEnabled(True)
        self.ui.export_tab.setEnabled(True)
        self.ui.plot_tab.setEnabled(True)

    def import_csv_file(self, preview=False, nrows=False, file_pathname=None):
        global status_bar_message, data, source_file_pathname

        if file_pathname is None:
            file_pathname = self.ui.filepathname_lineEdit.text()
            batch_mode = False
        else:
            batch_mode = True

        if file_pathname:
            source_file_pathname = file_pathname

            if not batch_mode:
                self.ui.export_data_lineEdit.setText(get_filename(source_file_pathname))
                self.init_on_import()

            delimiter = self.ui.import_csv_delimiter_comboBox.currentText()
            if delimiter == 'Auto':
                delimiter = None

            try:
                if self.ui.import_csv_skiprows_lineEdit.text():
                    skiprows = eval(self.ui.import_csv_skiprows_lineEdit.text())
                else:
                    skiprows = 0

                header_row = self.ui.import_csv_header_row_spinBox.value()

                if self.ui.import_csv_units_row_checkBox.isChecked():
                    units_row = self.ui.import_csv_units_row_spinBox.value()
                    skiprows += units_row - header_row
                else:
                    units_row = None

                try:
                    unitized_columns = get_unitized_columns(source_file_pathname,
                                                            encoding=self.ui.import_csv_encoding_comboBox.currentText(),
                                                            header_row=header_row, units_row=units_row,
                                                            delimiter=delimiter)

                    usecols = list(range(0, len(unitized_columns)))
                except:
                    unitized_columns = []
                    usecols = None

                keyword_args = self.import_csv_options_dict

                skip_blank_lines = eval(self.ui.import_csv_skip_blank_lines_comboBox.currentText())

                if preview:
                    try:
                        df = pd.read_csv(source_file_pathname, names=unitized_columns,
                                         header=header_row, delimiter=delimiter, usecols=usecols,
                                         encoding=self.ui.import_csv_encoding_comboBox.currentText(),
                                         skip_blank_lines=skip_blank_lines,
                                         nrows=nrows, skiprows=skiprows, on_bad_lines='warn',
                                         **keyword_args,
                                         )

                        df = self.handle_import_nans(df)
                    except:  # try without usecols...
                        try:
                            df = pd.read_csv(source_file_pathname, names=unitized_columns,
                                             header=header_row, delimiter=delimiter,  # usecols=usecols,
                                             encoding=self.ui.import_csv_encoding_comboBox.currentText(),
                                             skip_blank_lines=skip_blank_lines,
                                             nrows=nrows, skiprows=skiprows, on_bad_lines='warn',
                                             **keyword_args,
                                             )

                            df = self.handle_import_nans(df)
                        except:
                            df = None
                else:  # read in actual file
                    try:
                        df = pd.read_csv(source_file_pathname, names=unitized_columns, header=header_row,
                                         delimiter=delimiter, usecols=usecols,
                                         encoding=self.ui.import_csv_encoding_comboBox.currentText(),
                                         skip_blank_lines=skip_blank_lines, on_bad_lines='warn',
                                         skiprows=skiprows,
                                         **keyword_args,
                                         )

                        df = self.handle_import_nans(df)
                    except:  # try without usecols...
                        df = pd.read_csv(source_file_pathname, names=unitized_columns, header=header_row,
                                         delimiter=delimiter,  # usecols=usecols,
                                         encoding=self.ui.import_csv_encoding_comboBox.currentText(),
                                         skip_blank_lines=skip_blank_lines, on_bad_lines='warn',
                                         skiprows=skiprows,
                                         **keyword_args,
                                         )

                        df = self.handle_import_nans(df)

                    if not batch_mode:
                        data = df

                        self.setup_initial_triage_lists()

                        self.ui.statusbar.showMessage('imported %d rows of data, %d columns' %
                                                      (len(df), len(df.columns)), 10000)

                        self.ui.script_preview_consoleWidget.locals().update(globals())
                        self.ui.script_preview_consoleWidget.output.setPlainText('use run() to run user script!\n')

                return df

            except Exception as e:
                if not batch_mode:
                    QMessageBox(QMessageBox.Icon.Critical, 'CSV Import Error', 'Error reading "%s"\n\n%s' %
                                (file_pathname, str(e))).exec()
                else:
                    print('Excel Import Error')
                    print('Error reading "%s"\n\n%s' % (file_pathname, str(e)))

                return None

    def import_excel_file(self, preview=False, nrows=False, file_pathname=None):
        global status_bar_message, data, source_file_pathname

        if file_pathname is None:
            file_pathname = self.ui.filepathname_lineEdit.text()
            batch_mode = False
        else:
            batch_mode = True

        if file_pathname:
            source_file_pathname = file_pathname

            if not batch_mode:
                self.ui.export_data_lineEdit.setText(os.path.basename(file_pathname).rsplit('.', 1)[0])
                self.init_on_import()

            if self.ui.import_excel_sheet_comboBox.count() == 0:
                from openpyxl import load_workbook
                wb = load_workbook(filename=self.ui.filepathname_lineEdit.text(), read_only=True)
                self.ui.import_excel_sheet_comboBox.addItems(wb.sheetnames)
                self.ui.import_excel_sheet_comboBox.setCurrentIndex(0)

            sheet_name = self.ui.import_excel_sheet_comboBox.currentText()
            sheet_number = self.ui.import_excel_sheet_comboBox.currentIndex()

            header_row = self.ui.import_excel_header_row_spinBox.value()

            if self.ui.import_excel_units_row_checkBox.isChecked():
                units_row = self.ui.import_excel_units_row_spinBox.value()
            else:
                units_row = None

            try:
                if nrows is False:
                    nrows = None

                try:
                    # try loading by sheet name
                    unitized_columns = get_unitized_columns(source_file_pathname, sheet_name=sheet_name,
                                                            header_row=header_row, units_row=units_row)
                except:
                    # try loading by sheet number, in case sheet name changes across files...
                    unitized_columns = get_unitized_columns(source_file_pathname, sheet_name=sheet_number,
                                                            header_row=header_row, units_row=units_row)

                keyword_args = self.import_excel_options_dict

                engine_kwargs = {'read_only': True}

                if self.ui.import_excel_skiprows_lineEdit.text():
                    skiprows = eval(self.ui.import_excel_skiprows_lineEdit.text())
                else:
                    skiprows = None

                try:
                    # try loading by sheet name
                    df = pd.read_excel(source_file_pathname, names=unitized_columns, sheet_name=sheet_name,
                                       header=header_row, nrows=nrows, engine_kwargs=engine_kwargs, skiprows=skiprows,
                                       **keyword_args)
                except:
                    # try loading by sheet number, in case sheet name changes across files...
                    df = pd.read_excel(source_file_pathname, names=unitized_columns, sheet_name=sheet_number,
                                       header=header_row, nrows=nrows, engine_kwargs=engine_kwargs, skiprows=skiprows,
                                       **keyword_args)

                df = self.handle_import_nans(df)

                if not batch_mode:
                    self.ui.statusbar.showMessage('imported %d rows of df, %d columns' %
                                                  (len(df), len(df.columns)), 10000)

                    self.ui.script_preview_consoleWidget.locals().update(globals())
                    self.ui.script_preview_consoleWidget.output.setPlainText('use run() to run user script!\n')

                    if not preview:
                        data = df
                        self.setup_initial_triage_lists()

                return df

            except Exception as e:
                if not batch_mode:
                    QMessageBox(QMessageBox.Icon.Critical, 'Excel Import Error', 'Error reading "%s"\n\n%s' %
                                (file_pathname, str(e))).exec()
                else:
                    print('Excel Import Error')
                    print('Error reading "%s"\n\n%s' % (file_pathname, str(e)))
                return None

    def init_on_import(self):
        original_numeric_fields.clear()
        original_string_fields.clear()
        active_numeric_fields.clear()
        active_string_fields.clear()
        ignore_fields.clear()
        favorite_fields.clear()

        self.ui.triage_filter_widget.inputChanged()  # update triage widgets

        self.init_plot_widget(self.ui.graphic_preview_plot_widget, 'Graphic Preview')

    def script_run(self):
        global original_numeric_fields, original_string_fields, active_numeric_fields, active_string_fields, \
            ignore_fields

        self.ui.script_preview_consoleWidget.repl.runCmd('run()')
        # update consoleWidget namespace in case new vars created:
        self.ui.script_preview_consoleWidget.locals().update(globals())
        # update triage lists in case new fields created:

        original_numeric_fields = [c for c in data.select_dtypes(exclude=['string', 'object']).columns]
        original_string_fields = [c for c in data.select_dtypes(include='string').columns]
        original_object_fields = [c for c in data.select_dtypes(exclude=['object']).columns]

        non_ignored_numeric_fields = [f for f in original_numeric_fields if f not in ignore_fields]
        non_ignored_string_fields = [f for f in original_string_fields if f not in ignore_fields]

        active_numeric_fields.extend(non_ignored_numeric_fields)
        active_string_fields.extend(non_ignored_string_fields)

        active_numeric_fields = list(np.unique(active_numeric_fields))
        self.ui.triage_numeric_listWidget.source_data = active_numeric_fields

        active_string_fields = list(np.unique(active_string_fields))
        self.ui.triage_string_listWidget.source_data = active_string_fields

        self.ui.triage_filter_widget.inputChanged()  # update triage widgets

    def delete_ignores(self):
        global data, latest_item

        if self.ui.triage_ignore_listWidget.source_data:
            qb = QMessageBox()
            qb.setText('This action will permanently delete the data associated with these fields.')
            qb.setInformativeText('Data will have to be re-imported to recover after deletion.')
            qb.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            response = qb.exec()

            if response == QMessageBox.Ok:
                data.drop(columns=ignore_fields, inplace=True)
                ignore_fields.clear()
                latest_items.clear()
                latest_item = None
                self.ui.triage_filter_widget.inputChanged()  # update triage widgets

    def script_open(self):
        file_pathname = file_dialog('', '*.py', 'Load triage script')
        if file_pathname:
            with open(file_pathname, 'r') as f_read:
                self.ui.script_preview_plainTextEdit.setPlainText(f_read.read())

    def script_save(self):
        file_pathname = QFileDialog().getSaveFileName(self, 'Save triage script', os.getcwd(), '*.py', '*.py')[0]

        if file_pathname:
            with open(file_pathname, 'w') as f_write:
                f_write.write(self.ui.script_preview_plainTextEdit.toPlainText())

    def add_to_script(self):
        if latest_items:
            for i in latest_items:
                self.ui.script_preview_plainTextEdit.insertPlainText("data['%s']\n" % i)
        elif latest_item:
            self.ui.script_preview_plainTextEdit.insertPlainText("data['%s']\n" % latest_item)

        set_tab_by_name(self.ui.preview_tabWidget, 'Preprocess Script')

    def import_csv_freeform_changed(self):
        self.ui.import_csv_parameter_tableWidget.resizeColumnsToContents()

        self.import_csv_options_dict = (
            two_column_tableWidget_to_dict(self.ui.import_csv_parameter_tableWidget))

        self.load_file_preview()

    def import_excel_freeform_changed(self):
        self.ui.import_excel_parameter_tableWidget.resizeColumnsToContents()

        self.import_excel_options_dict = (
            two_column_tableWidget_to_dict(self.ui.import_excel_parameter_tableWidget))

        self.load_file_preview()

    def get_export_file_extension(self):
        if self.ui.export_data_comboBox.currentText() == 'CSV':
            file_extension = '.csv'
        else:
            file_extension = '.xlsx'

        return file_extension

    def export_file(self, export_data, file_name):

        if self.ui.export_all_but_ignored_radioButton.isChecked():
            export_data = export_data.drop(columns=ignore_fields, errors='ignore')
        elif self.ui.export_favorites_only_radioButton.isChecked():
            export_data = export_data.loc[:, favorite_fields]
        elif self.ui.export_numeric_only_radioButton.isChecked():
            export_data = export_data.select_dtypes(include='number')
        elif self.ui.export_non_numeric_only_radioButton.isChecked():
            export_data = export_data.select_dtypes(exclude='number')

        save_file_pathname, file_extension = (
            self.update_export_filename_preview(export_filename=file_name))

        print('exporting "%s"' % save_file_pathname)

        if file_extension == '.csv':
            export_data.to_csv(save_file_pathname, index=False)
        else:
            export_data.to_excel(save_file_pathname, index=False)

    def start_export_files_thread(self, source_files):
        self.ui.export_progressBar.setMaximum(0)
        self.export_worker_thread = QThread()
        self.export_worker = ExportWorker(source_files)
        self.export_worker.moveToThread(self.export_worker_thread)
        self.export_worker_thread.started.connect(self.export_worker.export_files)
        self.export_worker.progress.connect(self.update_progress)
        self.export_worker.finished.connect(self.export_worker_thread.quit)
        self.export_worker.finished.connect(self.export_worker.deleteLater)
        self.export_worker_thread.finished.connect(self.export_worker_thread.deleteLater)
        self.export_worker_thread.start()

    def cancel_export(self):
        self.export_worker.cancel()
        self.statusBar().showMessage('Cancelling Export...', 100000)

    def update_progress(self, value):
        global data

        self.ui.export_progressBar.setValue(value+1)

        if self.ui.export_progressBar.text() == '100%':
            if self.export_worker.cancelled:
                self.ui.export_progressBar.setValue(0)
                self.statusBar().showMessage('Export Cancelled!', 10000)
            else:
                self.statusBar().showMessage('Export Complete!', 10000)

            self.ui.export_options_groupBox.setEnabled(True)
            self.ui.export_data_cancel_pushButton.setEnabled(False)

        data = self.pre_export_data  # restore pre-export data

    def export_folder_browse(self):
        self.export_folder_pathname = QFileDialog().getExistingDirectory(self, 'Select Export Destination', '')

        self.ui.export_folder_filepathname_lineEdit.setText(self.export_folder_pathname)

        if self.export_folder_pathname:
            self.ui.export_data_pushButton.setEnabled(True)
            self.ui.export_mode_groupBox.setEnabled(True)
        else:
            self.ui.export_data_pushButton.setEnabled(False)
            self.ui.export_mode_groupBox.setEnabled(False)

    def select_export_batch_files(self):
        self.ui.export_batch_files_listWidget.clear()

        if get_current_tabname(self.ui.file_import_tabWidget) == 'CSV':
            filter_str = '*.csv'
        else:
            filter_str = '*.xls*'

        self.export_batch_source_files = QFileDialog().getOpenFileNames(self, 'Select Source Files', os.getcwd(),
                                                                        filter=filter_str, selectedFilter=filter_str)[0]
        if self.export_batch_source_files:
            self.ui.export_batch_files_listWidget.addItems(self.export_batch_source_files)
            self.ui.export_data_pushButton.setEnabled(True)
        else:
            self.ui.export_batch_files_listWidget.clear()
            self.ui.export_data_pushButton.setEnabled(False)

        self.ui.export_data_pushButton.setFocus()

    def remove_export_batch_files(self, event):
        if event.key() == Qt.Key.Key_Backspace or event.key() == Qt.Key.Key_Delete:
            selected_items = self.ui.export_batch_files_listWidget.selectedItems()
            if selected_items:
                for item in selected_items:
                    self.ui.export_batch_files_listWidget.takeItem(self.ui.export_batch_files_listWidget.row(item))
                    self.export_batch_source_files.remove(item.text())
        else:
            # If the key pressed is not Backspace or Delete, handle the event normally
            self.ui.export_batch_files_listWidget.default_keyPressEvent(event)

    def export_batch_files_listWidget_dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def export_batch_files_listWidget_dragMoveEvent(self, event):
        pass

    def export_batch_files_listWidget_dropEvent(self, event):
        if get_current_tabname(self.ui.file_import_tabWidget) == 'Excel':
            file_type = 'Excel'
        else:
            file_type = 'csv'

        for url in event.mimeData().urls():
            event.acceptProposedAction()

            file_path = url.toLocalFile()
            if file_type == 'csv' or (file_type == 'Excel' and 'xls' in file_path.split('.')[-1]):
                self.ui.export_batch_files_listWidget.addItem(file_path)
                self.export_batch_source_files.append(file_path)
                self.ui.export_data_pushButton.setEnabled(True)
            else:
                QMessageBox(QMessageBox.Icon.Critical,
                            'Drop Error', "Can't accept file type '%s' with current '%s' import settings" %
                            (file_path.split('.')[-1], file_type)).exec()

    def export_data(self):
        from file_io import create_combined_filename
        global data

        save_file_pathname = self.update_export_filename_preview()[0]

        if self.ui.export_mode_single_radioButton.isChecked():
            if data is not None:
                if self.export_folder_pathname:
                    self.pre_export_data = data.copy()

                    self.start_export_files_thread([self.input_file_pathname])

                    self.statusBar().showMessage('Exporting data to "%s" ...' % save_file_pathname, 100000)

                    self.ui.export_options_groupBox.setEnabled(False)
                    self.ui.export_data_cancel_pushButton.setEnabled(True)
        else:
            self.ui.export_progressBar.setValue(0)

            if self.export_batch_source_files:
                if self.export_folder_pathname:
                    self.pre_export_data = data.copy()

                    self.start_export_files_thread(self.export_batch_source_files)

                    self.statusBar().showMessage('Exporting %d files to "%s" ...' %
                                                 (len(self.export_batch_source_files), self.export_folder_pathname), 10000)

                    self.ui.export_options_groupBox.setEnabled(False)
                    self.ui.export_data_cancel_pushButton.setEnabled(True)

    def select_export_mode(self):
        self.ui.export_progressBar.setValue(0)

        if self.ui.export_mode_single_radioButton.isChecked():
            # self.ui.export_data_lineEdit.setEnabled(True)
            self.ui.export_mode_stackedWidget.setCurrentWidget(self.ui.single_page)
            self.ui.export_data_pushButton.setEnabled(True)

        if self.ui.export_mode_batch_radioButton.isChecked():
            self.ui.export_data_lineEdit.setEnabled(False)
            self.ui.export_mode_stackedWidget.setCurrentWidget(self.ui.batch_page)
            if self.ui.export_batch_files_listWidget.count() == 0:
                self.ui.export_data_pushButton.setEnabled(False)

    def update_export_filename_preview(self, dummyQString=None, export_filename=None):

        if self.ui.export_data_prefix_filler_comboBox.currentText() == 'None':
            prefix_filler = ''
        else:
            prefix_filler = self.ui.export_data_prefix_filler_comboBox.currentText()

        if self.ui.export_data_suffix_filler_comboBox.currentText() == 'None':
            suffix_filler = ''
        else:
            suffix_filler = self.ui.export_data_suffix_filler_comboBox.currentText()

        prefix = self.ui.export_data_prefix_lineEdit.text() + prefix_filler
        suffix = suffix_filler + self.ui.export_data_suffix_lineEdit.text()

        file_extension = self.get_export_file_extension()

        if export_filename is None:
            export_filename = self.ui.export_data_lineEdit.text()

        save_file_pathname = (
                self.export_folder_pathname + os.sep + prefix + get_filename(export_filename) + suffix + file_extension)

        self.ui.export_filename_preview_label.setText(save_file_pathname)

        return save_file_pathname, file_extension

    def closeEvent(self, event):
        self.active_plot.close()


def status_bar():
    print(time.time())
    if mainwindow:
        mainwindow.ui.statusbar.showMessage(status_bar_message, 1000)


def run_colliderscope():
    global app, mainwindow, timer
    import multitimer
    timer = multitimer.MultiTimer(interval=1, function=status_bar)
    app = QApplication(sys.argv)

    # set app style
    app.setStyle("Fusion")

    # customize tooltip style
    app.setStyleSheet("QToolTip {color: #ffffff; background-color: #402a82da; border: 2px solid black; "
                      "border-radius: 10px; padding: 2.5px; margin: 0px}")
    # app.setStyleSheet("QToolTip { color: #ffffff; background-color: #2a82da; opacity: 200; "
    #                   "border: 2px solid black; border-radius: 10px; padding: 2.5px; margin: 0px}")

    mainwindow = ColliderScopeUI()
    mainwindow.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    run_colliderscope()

# self.ui.graphic_preview_plot_widget.getViewBox().rbScaleBox.setPen(pg.mkPen((64, 128, 200), width=2)) #4080C8
# self.ui.graphic_preview_plot_widget.getViewBox().rbScaleBox.setBrush(pg.mkBrush(81, 197, 255, 100)) #51c5ff64
